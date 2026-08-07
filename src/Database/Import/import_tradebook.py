#!/usr/bin/env python3
"""Transactional, idempotent importer for the five authoritative Tradebook workbooks.

The importer is intentionally mapping-driven: workbook labels are untrusted input,
while database tables and columns must be declared in mapping.yaml and must exist in
the public schema. Values are always bound parameters and identifiers are composed
with psycopg.sql.Identifier.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
import yaml
from openpyxl.worksheet.worksheet import Worksheet

try:
    import psycopg
    from psycopg import sql
except ModuleNotFoundError:  # --validate-only and parser tests do not need a database driver.
    psycopg = None  # type: ignore[assignment]
    sql = None  # type: ignore[assignment]


IDENTIFIER = re.compile(r"^[a-z][a-z0-9_]*$")
SUPPORTED_TRANSFORMS = frozenset({"text", "decimal", "integer", "boolean", "date", "month"})
SYSTEM_ACTOR_ID = UUID("00000000-0000-0000-0000-000000000000")


class ImportContractError(ValueError):
    """The workbook or mapping does not satisfy the declared import contract."""


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().casefold())


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def transform(value: Any, kind: str, *, location: str) -> Any:
    if is_blank(value):
        return None
    if kind == "text":
        return str(value).strip()
    if kind == "decimal":
        try:
            if isinstance(value, str):
                value = value.strip().replace(" ", "").replace(",", ".")
            return Decimal(str(value))
        except (InvalidOperation, ValueError) as exc:
            raise ImportContractError(f"{location}: expected a decimal, received {value!r}") from exc
    if kind == "integer":
        try:
            converted = Decimal(str(value))
            if converted != converted.to_integral_value():
                raise ValueError
            return int(converted)
        except (InvalidOperation, ValueError) as exc:
            raise ImportContractError(f"{location}: expected an integer, received {value!r}") from exc
    if kind == "boolean":
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().casefold()
        if normalized in {"1", "true", "yes", "y", "active"}:
            return True
        if normalized in {"0", "false", "no", "n", "inactive"}:
            return False
        raise ImportContractError(f"{location}: expected a boolean, received {value!r}")
    if kind in {"date", "month"}:
        converted: date
        if isinstance(value, datetime):
            converted = value.date()
        elif isinstance(value, date):
            converted = value
        else:
            text = str(value).strip()
            converted = None  # type: ignore[assignment]
            for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m", "%m/%Y"):
                try:
                    converted = datetime.strptime(text, pattern).date()
                    break
                except ValueError:
                    continue
            if converted is None:
                raise ImportContractError(f"{location}: expected a date, received {value!r}")
        return converted.replace(day=1) if kind == "month" else converted
    raise ImportContractError(f"{location}: unsupported transform {kind!r}")


def _identifier(value: Any, location: str) -> str:
    if not isinstance(value, str) or not IDENTIFIER.fullmatch(value):
        raise ImportContractError(f"{location}: invalid SQL identifier {value!r}")
    return value


def load_manifest(path: Path) -> dict[str, Any]:
    try:
        manifest = yaml.safe_load(path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError) as exc:
        raise ImportContractError(f"Cannot read mapping manifest {path}: {exc}") from exc
    if not isinstance(manifest, dict) or manifest.get("version") != 1:
        raise ImportContractError("mapping.yaml must be an object with version: 1")

    workbooks = manifest.get("workbooks")
    mappings = manifest.get("mappings")
    allowed_tables = manifest.get("allowed_tables")
    if not isinstance(workbooks, dict) or not workbooks:
        raise ImportContractError("mapping.yaml must declare workbooks")
    if not isinstance(mappings, list) or not mappings:
        raise ImportContractError("mapping.yaml must declare mappings")
    if not isinstance(allowed_tables, list) or not allowed_tables:
        raise ImportContractError("mapping.yaml must declare allowed_tables")
    allowed = {_identifier(value, "allowed_tables") for value in allowed_tables}

    for index, item in enumerate(mappings):
        location = f"mappings[{index}]"
        if not isinstance(item, dict):
            raise ImportContractError(f"{location} must be an object")
        workbook = item.get("workbook")
        if workbook not in workbooks:
            raise ImportContractError(f"{location}.workbook references unknown workbook {workbook!r}")
        table = _identifier(item.get("table"), f"{location}.table")
        if table not in allowed:
            raise ImportContractError(f"{location}.table {table!r} is not allowlisted")
        sheets = item.get("sheets")
        if not isinstance(sheets, list) or not sheets or not all(isinstance(value, str) for value in sheets):
            raise ImportContractError(f"{location}.sheets must be a non-empty string list")
        keys = item.get("keys")
        columns = item.get("columns")
        if not isinstance(keys, list) or not keys:
            raise ImportContractError(f"{location}.keys must be a non-empty list")
        if not isinstance(columns, dict) or not columns:
            raise ImportContractError(f"{location}.columns must be an object")
        normalized_columns = {_identifier(name, f"{location}.columns") for name in columns}
        for key in keys:
            key = _identifier(key, f"{location}.keys")
            if key not in normalized_columns:
                raise ImportContractError(f"{location}: merge key {key!r} is not mapped")
        for target, column in columns.items():
            column_location = f"{location}.columns.{target}"
            if not isinstance(column, dict):
                raise ImportContractError(f"{column_location} must be an object")
            sources = column.get("sources", [])
            if "constant" not in column and (not isinstance(sources, list) or not sources):
                raise ImportContractError(f"{column_location} needs sources or constant")
            kind = column.get("transform", "text")
            if kind not in SUPPORTED_TRANSFORMS:
                raise ImportContractError(f"{column_location}: unsupported transform {kind!r}")
            lookup = column.get("lookup")
            if lookup is not None:
                if not isinstance(lookup, dict):
                    raise ImportContractError(f"{column_location}.lookup must be an object")
                for name in ("table", "key", "result"):
                    _identifier(lookup.get(name), f"{column_location}.lookup.{name}")
                if lookup["table"] not in allowed:
                    raise ImportContractError(f"{column_location}.lookup table is not allowlisted")
    return manifest


def parse_workbook_arguments(values: Sequence[str]) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for value in values:
        key, separator, raw_path = value.partition("=")
        if not separator or not key or not raw_path:
            raise ImportContractError(f"Invalid --workbook {value!r}; expected NAME=PATH")
        if key in result:
            raise ImportContractError(f"Workbook {key!r} was supplied more than once")
        path = Path(raw_path).expanduser().resolve()
        if not path.is_file():
            raise ImportContractError(f"Workbook {key!r} does not exist: {path}")
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ImportContractError(f"Workbook {key!r} must be .xlsx or .xlsm: {path}")
        result[key] = path
    return result


def select_sheet(workbook: openpyxl.Workbook, names: Sequence[str], location: str) -> Worksheet:
    by_normalized_name = {normalize_header(name): name for name in workbook.sheetnames}
    for candidate in names:
        actual = by_normalized_name.get(normalize_header(candidate))
        if actual is not None:
            return workbook[actual]
    raise ImportContractError(
        f"{location}: none of the configured sheets {list(names)!r} exist; found {workbook.sheetnames!r}")


def discover_header(
    worksheet: Worksheet,
    columns: Mapping[str, Mapping[str, Any]],
    *,
    location: str,
    max_header_row: int = 50,
) -> tuple[int, dict[str, int]]:
    best: tuple[int, int, dict[str, int]] | None = None
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(max_header_row, worksheet.max_row), values_only=True), start=1
    ):
        positions: dict[str, int] = {}
        normalized_positions: dict[str, int] = {}
        for index, value in enumerate(row):
            normalized = normalize_header(value)
            if normalized:
                if normalized in normalized_positions:
                    raise ImportContractError(f"{location} row {row_number}: duplicate header {value!r}")
                normalized_positions[normalized] = index
        score = 0
        for target, spec in columns.items():
            if "constant" in spec:
                continue
            for alias in spec.get("sources", []):
                position = normalized_positions.get(normalize_header(alias))
                if position is not None:
                    positions[target] = position
                    score += 1
                    break
        required_count = sum(1 for spec in columns.values() if spec.get("required") and "constant" not in spec)
        required_found = sum(1 for target, spec in columns.items() if spec.get("required") and target in positions)
        candidate = (required_found, score, positions)
        if best is None or candidate[:2] > best[:2]:
            best = candidate
        if required_found == required_count and score > 0:
            return row_number, positions

    missing = [
        target for target, spec in columns.items()
        if spec.get("required") and "constant" not in spec and (best is None or target not in best[2])
    ]
    raise ImportContractError(f"{location}: could not find a header row; missing required mappings {missing!r}")


def extract_rows(
    worksheet: Worksheet,
    columns: Mapping[str, Mapping[str, Any]],
    *,
    location: str,
) -> list[dict[str, Any]]:
    header_row, positions = discover_header(worksheet, columns, location=location)
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        source_values = [row[position] if position < len(row) else None for position in positions.values()]
        if not source_values or all(is_blank(value) for value in source_values):
            continue
        record: dict[str, Any] = {}
        for target, spec in columns.items():
            if "constant" in spec:
                value = spec["constant"]
            elif target in positions:
                position = positions[target]
                value = row[position] if position < len(row) else None
            else:
                continue
            value = transform(value, spec.get("transform", "text"), location=f"{location} row {row_number} {target}")
            if value is None and "default" in spec:
                value = transform(spec["default"], spec.get("transform", "text"), location=f"{location} default {target}")
            if value is None and spec.get("required"):
                raise ImportContractError(f"{location} row {row_number}: required value {target!r} is blank")
            record[target] = value
        records.append(record)
    return records


def validate_database_contract(
    connection: psycopg.Connection[Any], table: str, columns: Sequence[str], keys: Sequence[str]
) -> tuple[bool, bool]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            """,
            (table,),
        )
        actual = {row[0] for row in cursor.fetchall()}
    missing = set(columns) - actual
    if missing:
        raise ImportContractError(f"Database table {table!r} is missing mapped columns {sorted(missing)!r}")
    if not set(keys) <= actual:
        raise ImportContractError(f"Database table {table!r} is missing merge keys")
    return "version" in actual, "updated_at" in actual


def resolve_lookups(
    connection: psycopg.Connection[Any],
    records: list[dict[str, Any]],
    columns: Mapping[str, Mapping[str, Any]],
    *,
    location: str,
) -> None:
    cache: dict[tuple[str, str, str, Any], Any] = {}
    with connection.cursor() as cursor:
        for row_number, record in enumerate(records, start=1):
            for target, spec in columns.items():
                lookup = spec.get("lookup")
                value = record.get(target)
                if lookup is None or value is None:
                    continue
                cache_key = (lookup["table"], lookup["key"], lookup["result"], value)
                if cache_key not in cache:
                    cursor.execute(
                        sql.SQL("SELECT {result} FROM {table} WHERE {key} = %s").format(
                            result=sql.Identifier(lookup["result"]),
                            table=sql.Identifier(lookup["table"]),
                            key=sql.Identifier(lookup["key"]),
                        ),
                        (value,),
                    )
                    matches = cursor.fetchall()
                    if len(matches) != 1:
                        raise ImportContractError(
                            f"{location} record {row_number}: lookup for {target!r}={value!r} "
                            f"returned {len(matches)} rows"
                        )
                    cache[cache_key] = matches[0][0]
                record[target] = cache[cache_key]


def resolve_vat_fallbacks(
    connection: psycopg.Connection[Any],
    records: list[dict[str, Any]],
    *,
    location: str,
) -> None:
    """Fill blank book VAT percentages from the contract counterparty's country rate."""
    pending = [
        (row_number, record)
        for row_number, record in enumerate(records, start=1)
        if record.get("vat_pct") is None
    ]
    if not pending:
        return

    rates_by_contract: dict[Any, Decimal] = {}
    with connection.cursor() as cursor:
        for row_number, record in pending:
            contract_id = record.get("contract_id")
            if contract_id is None:
                raise ImportContractError(
                    f"{location} record {row_number}: VAT fallback requires contract_id"
                )
            if contract_id not in rates_by_contract:
                cursor.execute(
                    """
                    SELECT cp.vat_applicable, cp.country_code
                    FROM contracts AS contract
                    JOIN counterparties AS cp ON cp.id = contract.counterparty_id
                    WHERE contract.id = %s
                    """,
                    (contract_id,),
                )
                counterparty = cursor.fetchone()
                if counterparty is None:
                    raise ImportContractError(
                        f"{location} record {row_number}: contract {contract_id!s} has no counterparty"
                    )

                vat_applicable, country_code = counterparty
                if not vat_applicable:
                    rates_by_contract[contract_id] = Decimal("0")
                else:
                    cursor.execute(
                        """
                        SELECT DISTINCT vat_rate
                        FROM companies
                        WHERE country_code = %s AND vat_rate IS NOT NULL
                        """,
                        (country_code,),
                    )
                    country_rates = [row[0] for row in cursor.fetchall()]
                    if len(country_rates) != 1:
                        raise ImportContractError(
                            f"{location} record {row_number}: country {country_code!r} has "
                            f"{len(country_rates)} distinct VAT rates; expected exactly one"
                        )
                    rates_by_contract[contract_id] = country_rates[0]

            record["vat_pct"] = rates_by_contract[contract_id]


def merge_records(
    connection: psycopg.Connection[Any],
    table: str,
    records: list[dict[str, Any]],
    keys: Sequence[str],
) -> int:
    if not records:
        return 0
    columns = list(records[0])
    if any(list(record) != columns for record in records):
        raise ImportContractError(f"{table}: mapped records do not have a stable column set")
    versioned, timestamped = validate_database_contract(connection, table, columns, keys)
    seen: set[tuple[Any, ...]] = set()
    for record in records:
        key = tuple(record[column] for column in keys)
        if any(value is None for value in key):
            raise ImportContractError(f"{table}: merge key contains NULL: {key!r}")
        if key in seen:
            raise ImportContractError(f"{table}: duplicate workbook merge key: {key!r}")
        seen.add(key)

    stage = f"import_{table}"
    assignments = [column for column in columns if column not in keys]
    with connection.cursor() as cursor:
        # CTAS preserves PostgreSQL column types but deliberately does not copy
        # target NOT NULL/unique constraints. This lets target BEFORE triggers
        # populate derived fields such as contract_instance_id during the merge.
        cursor.execute(
            sql.SQL(
                "CREATE TEMP TABLE {stage} ON COMMIT DROP AS "
                "SELECT {columns} FROM {target} WITH NO DATA"
            ).format(
                stage=sql.Identifier(stage),
                columns=sql.SQL(", ").join(map(sql.Identifier, columns)),
                target=sql.Identifier(table),
            )
        )
        cursor.executemany(
            sql.SQL("INSERT INTO {stage} ({columns}) VALUES ({values})").format(
                stage=sql.Identifier(stage),
                columns=sql.SQL(", ").join(map(sql.Identifier, columns)),
                values=sql.SQL(", ").join(sql.Placeholder() for _ in columns),
            ),
            [tuple(record[column] for column in columns) for record in records],
        )

        match = sql.SQL(" AND ").join(
            sql.SQL("target.{key} IS NOT DISTINCT FROM source.{key}").format(key=sql.Identifier(key))
            for key in keys
        )
        changes = sql.SQL(" OR ").join(
            sql.SQL("target.{column} IS DISTINCT FROM source.{column}").format(
                column=sql.Identifier(column)
            )
            for column in assignments
        )
        update_parts = [
            sql.SQL("{column} = source.{column}").format(column=sql.Identifier(column))
            for column in assignments
        ]
        if assignments and versioned:
            update_parts.append(sql.SQL("version = target.version + 1"))
        if assignments and timestamped:
            update_parts.append(sql.SQL("updated_at = clock_timestamp()"))

        statement = sql.SQL("MERGE INTO {target} AS target USING {stage} AS source ON {match} ").format(
            target=sql.Identifier(table), stage=sql.Identifier(stage), match=match
        )
        if assignments:
            statement += sql.SQL("WHEN MATCHED AND ({changes}) THEN UPDATE SET {updates} ").format(
                changes=changes,
                updates=sql.SQL(", ").join(update_parts),
            )
        statement += sql.SQL("WHEN NOT MATCHED THEN INSERT ({columns}) VALUES ({values})").format(
            columns=sql.SQL(", ").join(map(sql.Identifier, columns)),
            values=sql.SQL(", ").join(
                sql.SQL("source.{column}").format(column=sql.Identifier(column)) for column in columns
            ),
        )
        cursor.execute(statement)
        # Several workbook sheets intentionally merge into the same target table
        # inside one transaction (for example Physical/Certificates sourcing and
        # sales). ON COMMIT DROP is only a final safety net; release the fixed-name
        # stage now so the next mapping can reuse it.
        cursor.execute(
            sql.SQL("DROP TABLE {stage}").format(stage=sql.Identifier(stage))
        )
    return len(records)


def summarize_record_counts(
    extracted: Iterable[tuple[Mapping[str, Any], Sequence[Mapping[str, Any]], str]],
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for mapping, rows, _ in extracted:
        table = mapping["table"]
        counts[table] = counts.get(table, 0) + len(rows)
    return counts


def run_import(
    database_url: str,
    manifest: Mapping[str, Any],
    workbook_paths: Mapping[str, Path],
    *,
    validate_only: bool,
    actor_id: UUID,
) -> dict[str, int]:
    expected = set(manifest["workbooks"])
    supplied = set(workbook_paths)
    if supplied != expected:
        raise ImportContractError(
            f"Exactly the configured workbooks are required; missing={sorted(expected - supplied)!r}, "
            f"unknown={sorted(supplied - expected)!r}"
        )

    opened: dict[str, openpyxl.Workbook] = {}
    extracted: list[tuple[Mapping[str, Any], list[dict[str, Any]], str]] = []
    try:
        for key, path in workbook_paths.items():
            opened[key] = openpyxl.load_workbook(
                path, read_only=True, data_only=True, keep_vba=path.suffix.casefold() == ".xlsm"
            )
        for mapping in manifest["mappings"]:
            workbook = opened[mapping["workbook"]]
            location = f"{workbook_paths[mapping['workbook']].name}/{mapping['sheets'][0]}"
            worksheet = select_sheet(workbook, mapping["sheets"], location)
            rows = extract_rows(worksheet, mapping["columns"], location=location)
            extracted.append((mapping, rows, location))

        if validate_only:
            return summarize_record_counts(extracted)

        if psycopg is None or sql is None:
            raise ImportContractError(
                "Database import requires psycopg; install src/Database/Import/requirements.txt"
            )

        counts: dict[str, int] = {}
        with psycopg.connect(database_url) as connection:
            with connection.transaction():
                with connection.cursor() as cursor:
                    cursor.execute("SELECT pg_advisory_xact_lock(%s)", (781_240_011,))
                    cursor.execute("SELECT set_config('app.actor_id', %s, true)", (str(actor_id),))
                for mapping, rows, location in extracted:
                    resolve_lookups(connection, rows, mapping["columns"], location=location)
                    if "vat_pct" in mapping["columns"]:
                        resolve_vat_fallbacks(connection, rows, location=location)
                    count = merge_records(connection, mapping["table"], rows, mapping["keys"])
                    counts[mapping["table"]] = counts.get(mapping["table"], 0) + count
            return counts
    finally:
        for workbook in opened.values():
            workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database-url", help="PostgreSQL connection string; required unless --validate-only")
    parser.add_argument(
        "--mapping",
        type=Path,
        default=Path(__file__).with_name("mapping.yaml"),
        help="Path to the allowlisted workbook mapping manifest",
    )
    parser.add_argument(
        "--workbook",
        action="append",
        default=[],
        metavar="NAME=PATH",
        help="Workbook binding; provide each configured workbook exactly once",
    )
    parser.add_argument("--validate-only", action="store_true", help="Validate workbook sheets, headers, and values")
    parser.add_argument("--actor-id", type=UUID, default=SYSTEM_ACTOR_ID, help="Audit actor for this controlled import")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        if not args.validate_only and not args.database_url:
            raise ImportContractError("--database-url is required unless --validate-only is used")
        manifest = load_manifest(args.mapping.resolve())
        workbook_paths = parse_workbook_arguments(args.workbook)
        counts = run_import(
            args.database_url or "",
            manifest,
            workbook_paths,
            validate_only=args.validate_only,
            actor_id=args.actor_id,
        )
        for table, count in sorted(counts.items()):
            print(f"{table}: {count} validated row(s)" if args.validate_only else f"{table}: {count} merged row(s)")
        return 0
    except ImportContractError as exc:
        print(f"import contract error: {exc}", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"import failed and was rolled back: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        if psycopg is not None and isinstance(exc, psycopg.Error):
            print(f"import failed and was rolled back: {exc}", file=sys.stderr)
            return 1
        raise


if __name__ == "__main__":
    raise SystemExit(main())
